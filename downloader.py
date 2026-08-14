import re
import zipfile
import shutil
import requests
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN
# ============================================================

YEARS = range(2012, 2027)

VERIFY_SSL = True

BASE_URL = "https://contrataciondelsectorpublico.gob.es"

# Carpeta Downloads del usuario
DOWNLOADS_DIR = Path.home() / "Downloads"

# Carpeta raíz del proceso
WORK_DIR = DOWNLOADS_DIR / "PLACSP_AEAT"

# Carpeta donde se guardan los ZIPs descargados
ZIP_DIR = WORK_DIR / "zips_descargados"

# Carpeta donde se conservan los ficheros .atom / .xml extraídos
ATOM_DIR = WORK_DIR / "atom_extraidos"

# Excel final
OUTPUT_FILE = WORK_DIR / "licitaciones_aeat_2012_2026.xlsx"

PATRONES_ORGANO = [
    r"Agencia Estatal de Administración Tributaria",
    r"\bAEAT\b",
    r"Departamento de Informática Tributaria",
    r"Delegación Especial de la AEAT",
]

COLUMNAS = [
    "Año fuente",
    "Expediente",
    "Objeto",
    "Órgano de contratación",
    "Coincidencia filtro",
    "Estado",
    "Importe sin IVA",
    "Importe con IVA",
    "Tipo de contrato",
    "Procedimiento",
    "CPV",
    "Fecha publicación",
    "Fecha actualización",
    "URL",
]


# ============================================================
# CREAR CARPETAS
# ============================================================

WORK_DIR.mkdir(parents=True, exist_ok=True)
ZIP_DIR.mkdir(parents=True, exist_ok=True)
ATOM_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# NAMESPACES ATOM / CODICE
# ============================================================

NS = {
    "atom": "http://www.w3.org/2005/Atom",
    "cbc": "urn:dgpe:names:draft:codice:schema:xsd:CommonBasicComponents-2",
    "cac": "urn:dgpe:names:draft:codice:schema:xsd:CommonAggregateComponents-2",
    "cac-place-ext": "urn:dgpe:names:draft:codice-place-ext:schema:xsd:CommonAggregateComponents-2",
}


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def construir_url_zip(year):
    return (
        BASE_URL
        + f"/sindicacion/sindicacion_643/licitacionesPerfilesContratanteCompleto3_{year}.zip"
    )


def local_name(tag):
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def get_text(element, path):
    found = element.find(path, NS)
    if found is not None and found.text:
        return found.text.strip()
    return None


def get_first_text(element, paths):
    for path in paths:
        value = get_text(element, path)
        if value:
            return value
    return None


def convertir_importe(valor):
    """
    Convierte importes a float.

    Soporta formatos como:
    649777.1
    49.949,00 €
    55.348.960,00 €
    60438.29
    """
    if valor is None:
        return None

    texto = str(valor).strip()

    if texto == "":
        return None

    texto = texto.replace("€", "").replace(" ", "").strip()

    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto and "." not in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return None


def convertir_entero(valor):
    if valor is None:
        return None

    texto = str(valor).strip()

    if texto == "":
        return None

    try:
        return int(float(texto))
    except ValueError:
        return None


def convertir_fecha(valor):
    """
    Convierte timestamps tipo:
    2012-01-10T09:44:39.589+01:00
    2012-01-10
    a objeto date.

    En Excel se mostrará como DD-MM-AAAA.
    """
    if valor is None:
        return None

    texto = str(valor).strip()

    if texto == "":
        return None

    texto = texto.replace("Z", "+00:00")

    try:
        return datetime.fromisoformat(texto).date()
    except ValueError:
        pass

    try:
        return datetime.strptime(texto[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def obtener_url_entry(entry):
    for link_node in entry.findall(".//atom:link", NS):
        href = link_node.attrib.get("href")
        if href:
            return href

    return None


def obtener_coincidencia_organo(organo):
    if not organo:
        return None

    for patron in PATRONES_ORGANO:
        if re.search(patron, organo, re.IGNORECASE):
            return patron

    return None


# ============================================================
# FECHAS
# ============================================================

def obtener_fecha_actualizacion(entry):
    valor = get_first_text(entry, [
        ".//atom:updated"
    ])

    return convertir_fecha(valor)


def obtener_fecha_publicacion(entry):
    """
    La fecha de publicación no siempre aparece en atom:published.
    Por eso se buscan nodos CODICE típicos de publicación.

    Si no se encuentra nada, se usa la fecha de actualización como fallback,
    para evitar dejar la columna vacía.
    """

    paths_prioritarios = [
        ".//cac-place-ext:ValidNoticeInfo//cbc:IssueDate",
        ".//cac-place-ext:ValidNoticeInfo//cbc:PublicationDate",
        ".//cac-place-ext:AdditionalPublicationStatus//cbc:IssueDate",
        ".//cac-place-ext:AdditionalPublicationStatus//cbc:PublicationDate",
        ".//cac:AdditionalDocumentReference/cbc:IssueDate",
        ".//cbc:IssueDate",
        ".//cbc:PublicationDate",
        ".//atom:published",
    ]

    valor = get_first_text(entry, paths_prioritarios)

    if valor:
        fecha = convertir_fecha(valor)
        if fecha:
            return fecha

    nombres_fecha_publicacion = {
        "IssueDate",
        "PublicationDate",
        "PublishedDate",
        "NoticePublicationDate",
    }

    for elem in entry.iter():
        nombre = local_name(elem.tag)

        if nombre in nombres_fecha_publicacion and elem.text:
            fecha = convertir_fecha(elem.text)
            if fecha:
                return fecha

    # Fallback final
    return obtener_fecha_actualizacion(entry)


# ============================================================
# PARSEO DE ENTRY
# ============================================================

def parse_entry(entry, year):
    expediente = get_first_text(entry, [
        ".//cbc:ContractFolderID"
    ])

    objeto = get_first_text(entry, [
        ".//cac:ProcurementProject/cbc:Name",
        ".//cbc:Title",
        ".//atom:title"
    ])

    organo = get_first_text(entry, [
        ".//cac-place-ext:ContractFolderStatus/cac-place-ext:LocatedContractingParty/cac:PartyName/cbc:Name",
        ".//cac:ContractingParty/cac:Party/cac:PartyName/cbc:Name",
        ".//cac:PartyName/cbc:Name",
    ])

    estado = get_first_text(entry, [
        ".//cbc:ContractFolderStatusCode"
    ])

    importe_sin_iva = get_first_text(entry, [
        ".//cac:ProcurementProject/cac:BudgetAmount/cbc:TaxExclusiveAmount"
    ])

    importe_con_iva = get_first_text(entry, [
        ".//cac:ProcurementProject/cac:BudgetAmount/cbc:TotalAmount"
    ])

    tipo_contrato = get_first_text(entry, [
        ".//cac:ProcurementProject/cbc:TypeCode"
    ])

    procedimiento = get_first_text(entry, [
        ".//cac:TenderingProcess/cbc:ProcedureCode"
    ])

    cpv = get_first_text(entry, [
        ".//cac:RequiredCommodityClassification/cbc:ItemClassificationCode"
    ])

    return {
        "Año fuente": int(year),
        "Expediente": expediente,
        "Objeto": objeto,
        "Órgano de contratación": organo,
        "Estado": estado,
        "Importe sin IVA": convertir_importe(importe_sin_iva),
        "Importe con IVA": convertir_importe(importe_con_iva),
        "Tipo de contrato": convertir_entero(tipo_contrato),
        "Procedimiento": convertir_entero(procedimiento),
        "CPV": str(cpv).strip() if cpv else None,
        "Fecha publicación": obtener_fecha_publicacion(entry),
        "Fecha actualización": obtener_fecha_actualizacion(entry),
        "URL": obtener_url_entry(entry),
    }


# ============================================================
# DESCARGA DE ZIP
# ============================================================

def descargar_zip(year):
    url = construir_url_zip(year)
    zip_path = ZIP_DIR / f"licitaciones_{year}.zip"

    if zip_path.exists() and zip_path.stat().st_size > 0:
        print(f"[{year}] ZIP ya existe. Se reutiliza:")
        print(f"[{year}] {zip_path}")
        return zip_path

    print(f"[{year}] Descargando ZIP...")
    print(f"[{year}] {url}")

    try:
        with requests.get(
            url,
            stream=True,
            timeout=180,
            verify=VERIFY_SSL
        ) as response:

            if response.status_code == 404:
                print(f"[{year}] No existe ZIP. Saltando.")
                return None

            response.raise_for_status()

            with open(zip_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        f.write(chunk)

        print(f"[{year}] ZIP descargado:")
        print(f"[{year}] {zip_path}")
        return zip_path

    except Exception as e:
        print(f"[{year}] Error descargando ZIP: {e}")

        if zip_path.exists():
            zip_path.unlink()

        return None


# ============================================================
# EXTRACCIÓN DE ATOM
# ============================================================

def extraer_atom(year, zip_path):
    year_atom_dir = ATOM_DIR / str(year)
    year_atom_dir.mkdir(parents=True, exist_ok=True)

    atom_files_existentes = (
        list(year_atom_dir.glob("*.atom"))
        + list(year_atom_dir.glob("*.xml"))
    )

    if atom_files_existentes:
        print(f"[{year}] ATOM ya extraídos. Se reutilizan {len(atom_files_existentes)} ficheros.")
        return atom_files_existentes

    print(f"[{year}] Extrayendo ficheros ATOM en:")
    print(f"[{year}] {year_atom_dir}")

    atom_paths = []

    with zipfile.ZipFile(zip_path, "r") as z:
        for member in z.namelist():
            lower = member.lower()

            if not (lower.endswith(".atom") or lower.endswith(".xml")):
                continue

            original_name = Path(member).name
            output_path = year_atom_dir / original_name

            counter = 1
            while output_path.exists():
                output_path = year_atom_dir / f"{Path(original_name).stem}_{counter}{Path(original_name).suffix}"
                counter += 1

            with z.open(member) as source, open(output_path, "wb") as target:
                shutil.copyfileobj(source, target)

            atom_paths.append(output_path)

    print(f"[{year}] ATOM extraídos: {len(atom_paths)}")
    return atom_paths


# ============================================================
# PROCESADO DE ATOM POR AÑO
# ============================================================

def procesar_year(year):
    registros = []

    zip_path = descargar_zip(year)

    if zip_path is None:
        return registros

    atom_paths = extraer_atom(year, zip_path)

    print(f"[{year}] Procesando ATOM...")

    for atom_path in atom_paths:
        try:
            tree = ET.parse(atom_path)
            root = tree.getroot()

            entries = root.findall(".//atom:entry", NS)

            for entry in entries:
                data = parse_entry(entry, year)

                organo = data.get("Órgano de contratación") or ""
                coincidencia = obtener_coincidencia_organo(organo)

                if coincidencia:
                    data["Coincidencia filtro"] = coincidencia
                    registros.append(data)

            root.clear()

        except ET.ParseError:
            print(f"[{year}] Error XML en {atom_path.name}. Saltando.")
            continue

        except Exception as e:
            print(f"[{year}] Error procesando {atom_path.name}: {e}")
            continue

    print(f"[{year}] Registros encontrados: {len(registros)}")
    return registros


# ============================================================
# EXCEL
# ============================================================

def ajustar_hoja(ws, registros, year):
    ws.append(COLUMNAS)
    ws.freeze_panes = "A2"

    for row in registros:
        ws.append([row.get(col) for col in COLUMNAS])

    max_row = ws.max_row
    max_col = ws.max_column

    # Formatos
    for row_num in range(2, max_row + 1):
        # Año
        ws[f"A{row_num}"].number_format = "0"

        # Importes
        ws[f"G{row_num}"].number_format = '#,##0.00 €'
        ws[f"H{row_num}"].number_format = '#,##0.00 €'

        # Códigos numéricos
        ws[f"I{row_num}"].number_format = "0"
        ws[f"J{row_num}"].number_format = "0"

        # CPV como texto
        ws[f"K{row_num}"].number_format = "@"

        # Fechas
        ws[f"L{row_num}"].number_format = "DD-MM-YYYY"
        ws[f"M{row_num}"].number_format = "DD-MM-YYYY"

    # Crear tabla Excel si hay registros
    if max_row >= 2:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"

        table = Table(
            displayName=f"Tabla_{year}",
            ref=ref
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style
        ws.add_table(table)

    # Anchuras de columnas
    anchos = {
        "A": 12,
        "B": 22,
        "C": 70,
        "D": 50,
        "E": 35,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 14,
        "L": 18,
        "M": 18,
        "N": 90,
    }

    for col, width in anchos.items():
        ws.column_dimensions[col].width = width


def crear_hoja_o_reemplazar(wb, year, registros):
    sheet_name = str(year)

    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    ws = wb.create_sheet(title=sheet_name)
    ajustar_hoja(ws, registros, year)


def crear_workbook_vacio():
    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)
    return wb


# ============================================================
# MAIN
# ============================================================

def main():
    print("")
    print("=================================")
    print(" PLACSP - AEAT")
    print("=================================")
    print("")
    print(f"Carpeta de salida: {WORK_DIR}")
    print(f"Carpeta ZIPs: {ZIP_DIR}")
    print(f"Carpeta ATOM: {ATOM_DIR}")
    print(f"Excel: {OUTPUT_FILE}")
    print(f"Años: {list(YEARS)}")
    print("")

    wb = crear_workbook_vacio()

    total = 0

    for year in YEARS:
        print("")
        print("=================================")
        print(f" Procesando año {year}")
        print("=================================")

        registros = procesar_year(year)

        crear_hoja_o_reemplazar(wb, year, registros)

        total += len(registros)

        print(f"[{year}] Guardando Excel actualizado...")
        wb.save(OUTPUT_FILE)
        print(f"[{year}] Excel actualizado en:")
        print(OUTPUT_FILE)

    print("")
    print("=================================")
    print(" PROCESO FINALIZADO")
    print("=================================")
    print(f"Total registros encontrados: {total}")
    print(f"Excel generado en: {OUTPUT_FILE}")
    print(f"ZIPs conservados en: {ZIP_DIR}")
    print(f"ATOM conservados en: {ATOM_DIR}")


if __name__ == "__main__":
    main()
